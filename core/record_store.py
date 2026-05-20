from __future__ import annotations

import json
import os
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import requests
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook


OUTPUT_DIR = Path("outputs")
RECORD_XLSX = OUTPUT_DIR / "research_screening_records.xlsx"
DEFAULT_SHEET_ID = "1pwuRill-lgd_HDCD1wz7J32PPQviKXmiq8F_9NkUSzA"
PRIVACY_NOTICE_VERSION = "uk-gdpr-research-v1"

EMAIL_RE = re.compile(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}")
PHONE_RE = re.compile(r"(?:\+?\d[\d\s().-]{7,}\d)")
URL_RE = re.compile(r"https?://\S+|www\.\S+", re.IGNORECASE)


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def _env_bool(name: str, default: bool = False) -> bool:
    value = os.getenv(name)
    if value is None:
        return default
    return str(value).strip().lower() in {"1", "true", "yes", "on"}


def _load_env() -> None:
    load_dotenv(Path(".env"), override=False)


def sanitize_free_text(text: str | None, max_len: int = 500) -> str:
    if not text:
        return ""
    value = str(text)
    value = EMAIL_RE.sub("[redacted-email]", value)
    value = PHONE_RE.sub("[redacted-phone]", value)
    value = URL_RE.sub("[redacted-url]", value)
    value = " ".join(value.split())
    return value[:max_len]


def as_json(value: Any) -> str:
    try:
        return json.dumps(value or {}, ensure_ascii=False, separators=(",", ":"))
    except Exception:
        return "{}"


def _top_interview_themes(answers: list[str]) -> tuple[str, str]:
    theme_tokens = {
        "sleep": {"sleep", "insomnia", "tired", "fatigue", "night", "rest"},
        "work_study": {"work", "job", "study", "college", "university", "deadline", "focus"},
        "relationships": {"family", "friend", "relationship", "alone", "lonely", "partner"},
        "health": {"pain", "body", "heart", "breath", "panic", "appetite", "health"},
    }
    text = " ".join(answers).lower()
    scores = {
        theme: sum(1 for token in tokens if token in text)
        for theme, tokens in theme_tokens.items()
    }
    ranked = sorted(scores.items(), key=lambda item: item[1], reverse=True)
    primary = ranked[0][0] if ranked and ranked[0][1] > 0 else ""
    secondary = ranked[1][0] if len(ranked) > 1 and ranked[1][1] > 0 else ""
    return primary, secondary


def get_record_headers() -> list[str]:
    return [
        "session_id",
        "record_id",
        "app_version",
        "record_created_at_utc",
        "record_updated_at_utc",
        "screening_completed_at_utc",
        "feedback_completed_at_utc",
        "deployment_environment",
        "consent_completed",
        "consent_research_data_storage",
        "consent_google_sheet_storage",
        "consent_pdf_report",
        "privacy_notice_version",
        "consent_timestamp_utc",
        "age_group",
        "gender_option",
        "preferred_language",
        "device_type",
        "previous_screening_experience",
        "accessibility_support_needed",
        "interview_complete",
        "interview_answer_count",
        "interview_theme_primary",
        "interview_theme_secondary",
        "emotion_label",
        "emotion_confidence",
        "emotion_scores_json",
        "risk_level",
        "risk_flags_json",
        "safety_flag_present",
        "bcqol_complete",
        "bcqol_raw_total",
        "bcqol_max_total",
        "bcqol_normalised_score",
        "bcqol_category",
        "mmse_complete",
        "mmse_total",
        "mmse_max_total",
        "mmse_orientation_score",
        "mmse_registration_score",
        "mmse_attention_score",
        "mmse_recall_score",
        "mmse_language_score",
        "mmse_category",
        "mmse_accessibility_flags_json",
        "combined_screening_status",
        "outcome_label",
        "recommendation_level",
        "report_generated",
        "pdf_generated",
        "feedback_submitted",
        "feedback_ease_of_use",
        "feedback_question_clarity",
        "feedback_report_clarity",
        "feedback_comfort",
        "feedback_trust",
        "feedback_usefulness",
        "feedback_recommend",
        "feedback_open_comment_redacted",
        "missing_sections_json",
        "validation_warnings_json",
        "record_status",
        "save_target_google_success",
        "save_target_local_success",
    ]


def get_audit_headers() -> list[str]:
    return ["timestamp_utc", "session_id", "save_target", "success", "error_message"]


def _normalise_row(record: dict[str, Any], headers: list[str]) -> list[Any]:
    return [record.get(header, "") for header in headers]


def _get_existing_value(session_id: str, column: str, worksheet_name: str) -> Any:
    if not RECORD_XLSX.exists():
        return ""
    try:
        workbook = load_workbook(RECORD_XLSX)
        if worksheet_name not in workbook.sheetnames:
            workbook.close()
            return ""
        worksheet = workbook[worksheet_name]
        headers = [worksheet.cell(row=1, column=index).value for index in range(1, worksheet.max_column + 1)]
        if column not in headers:
            workbook.close()
            return ""
        column_index = headers.index(column) + 1
        for row_index in range(2, worksheet.max_row + 1):
            if str(worksheet.cell(row=row_index, column=1).value) == str(session_id):
                value = worksheet.cell(row=row_index, column=column_index).value
                workbook.close()
                return value or ""
        workbook.close()
    except Exception:
        return ""
    return ""


def _derive_outcome(session_state: dict, missing_sections: list[str], validation_warnings: list[str]) -> tuple[str, str]:
    if missing_sections:
        return "Incomplete assessment", "complete missing sections"

    emotion_result = session_state.get("emotion_result") or {}
    risk_flags = emotion_result.get("risk_flags", {})
    if session_state.get("safety_stop") or risk_flags.get("self_harm_ideation"):
        return "Safety signposting triggered", "urgent safety signposting"

    mmse_result = session_state.get("mmse_result") or {}
    if mmse_result.get("risk_flag") or str(mmse_result.get("interpretation_detail", {}).get("severity", "")).endswith("_flag"):
        return "Cognitive screening concerns indicated", "consider speaking to a professional"

    bcqol_result = session_state.get("bcqol_result") or {}
    bcqol_score = bcqol_result.get("score_total")
    moderate_emotion_risk = risk_flags.get("high_distress") or risk_flags.get("needs_human_followup")
    if (isinstance(bcqol_score, (int, float)) and bcqol_score < 50) or moderate_emotion_risk:
        return "Some wellbeing concerns indicated", "consider speaking to a professional"

    if validation_warnings:
        return "Some wellbeing concerns indicated", "consider speaking to a professional"
    return "No immediate concern indicated", "self-care information"


def build_screening_record(session_state: dict) -> dict[str, Any]:
    _load_env()
    session_id = session_state.get("session_id", "")
    consent = session_state.get("consent") or {}
    demographics = consent.get("demographics", {})
    confirmed = consent.get("confirmed_items", {})
    emotion_result = session_state.get("emotion_result") or {}
    risk_flags = emotion_result.get("risk_flags", {})
    bcqol_result = session_state.get("bcqol_result") or {}
    mmse_result = session_state.get("mmse_result") or {}
    domain_scores = mmse_result.get("domain_scores", {})
    answers = session_state.get("answers", [])
    primary_theme, secondary_theme = _top_interview_themes(answers)

    missing_sections: list[str] = []
    if not consent.get("is_consented"):
        missing_sections.append("consent")
    if not session_state.get("emotion_result"):
        missing_sections.append("interview_emotion_summary")
    if not session_state.get("bcqol_result"):
        missing_sections.append("bcqol")
    if not session_state.get("mmse_result"):
        missing_sections.append("mmse")

    validation_warnings = list(mmse_result.get("review_flags", []))
    if risk_flags.get("high_distress") or risk_flags.get("needs_human_followup"):
        validation_warnings.append("emotion_followup_flag")
    outcome_label, recommendation_level = _derive_outcome(session_state, missing_sections, validation_warnings)

    emotion_probs = emotion_result.get("emotion_probs", {})
    emotion_confidence = max(emotion_probs.values(), default=0.0) if isinstance(emotion_probs, dict) else ""
    risk_level = "urgent" if risk_flags.get("self_harm_ideation") else "moderate" if (
        risk_flags.get("high_distress") or risk_flags.get("needs_human_followup") or mmse_result.get("risk_flag")
    ) else "low"
    now = utc_now()
    record_created = _get_existing_value(session_id, "record_created_at_utc", "screening_records") or session_state.get("created_utc", now)

    return {
        "session_id": session_id,
        "record_id": f"research_{session_id}",
        "app_version": os.getenv("APP_VERSION", "prototype-feedback-records-v1"),
        "record_created_at_utc": record_created,
        "record_updated_at_utc": now,
        "screening_completed_at_utc": session_state.get("screening_completed_at_utc") or now,
        "feedback_completed_at_utc": session_state.get("feedback_completed_at_utc", ""),
        "deployment_environment": os.getenv("DEPLOYMENT_ENVIRONMENT", "local_or_streamlit"),
        "consent_completed": bool(consent.get("is_consented")),
        "consent_research_data_storage": bool(confirmed.get("research_data_storage")),
        "consent_google_sheet_storage": bool(confirmed.get("google_sheet_storage")),
        "consent_pdf_report": True,
        "privacy_notice_version": os.getenv("RESEARCH_PRIVACY_NOTICE_VERSION", PRIVACY_NOTICE_VERSION),
        "consent_timestamp_utc": consent.get("timestamp", ""),
        "age_group": demographics.get("age_group", ""),
        "gender_option": demographics.get("gender", ""),
        "preferred_language": demographics.get("preferred_language", ""),
        "device_type": demographics.get("device_type", ""),
        "previous_screening_experience": demographics.get("previous_screening_experience", ""),
        "accessibility_support_needed": bool(demographics.get("accessibility_support_required")),
        "interview_complete": len(answers) >= 5,
        "interview_answer_count": len(answers),
        "interview_theme_primary": primary_theme,
        "interview_theme_secondary": secondary_theme,
        "emotion_label": emotion_result.get("emotion_label", ""),
        "emotion_confidence": emotion_confidence,
        "emotion_scores_json": as_json(emotion_probs),
        "risk_level": risk_level,
        "risk_flags_json": as_json(risk_flags),
        "safety_flag_present": bool(session_state.get("safety_stop") or risk_flags.get("self_harm_ideation")),
        "bcqol_complete": bool(bcqol_result),
        "bcqol_raw_total": bcqol_result.get("raw_total", ""),
        "bcqol_max_total": bcqol_result.get("max_possible_score", ""),
        "bcqol_normalised_score": bcqol_result.get("score_total", ""),
        "bcqol_category": bcqol_result.get("interpretation", ""),
        "mmse_complete": bool(mmse_result.get("success")),
        "mmse_total": mmse_result.get("total_score", ""),
        "mmse_max_total": mmse_result.get("max_score", ""),
        "mmse_orientation_score": domain_scores.get("orientation_time", 0) + domain_scores.get("orientation_place", 0) if domain_scores else "",
        "mmse_registration_score": domain_scores.get("registration", ""),
        "mmse_attention_score": domain_scores.get("attention_calculation", ""),
        "mmse_recall_score": domain_scores.get("recall", ""),
        "mmse_language_score": domain_scores.get("language", ""),
        "mmse_category": mmse_result.get("interpretation", ""),
        "mmse_accessibility_flags_json": as_json((session_state.get("mmse_form_state") or {}).get("accessibility_flags", {})),
        "combined_screening_status": "complete" if not missing_sections else "incomplete",
        "outcome_label": outcome_label,
        "recommendation_level": recommendation_level,
        "report_generated": bool(session_state.get("report_json")),
        "pdf_generated": bool(session_state.get("pdf_bytes")),
        "feedback_submitted": bool(session_state.get("feedback_submitted")),
        "feedback_ease_of_use": session_state.get("feedback_ease_of_use", ""),
        "feedback_question_clarity": session_state.get("feedback_question_clarity", ""),
        "feedback_report_clarity": session_state.get("feedback_report_clarity", ""),
        "feedback_comfort": session_state.get("feedback_comfort", ""),
        "feedback_trust": session_state.get("feedback_trust", ""),
        "feedback_usefulness": session_state.get("feedback_usefulness", ""),
        "feedback_recommend": session_state.get("feedback_recommend", ""),
        "feedback_open_comment_redacted": sanitize_free_text(session_state.get("feedback_open_comment_redacted", "")),
        "missing_sections_json": as_json(missing_sections),
        "validation_warnings_json": as_json(sorted(set(validation_warnings))),
        "record_status": "screening_saved",
        "save_target_google_success": False,
        "save_target_local_success": False,
    }


def build_feedback_record(session_state: dict) -> dict[str, Any]:
    record = build_screening_record(session_state)
    completed_at = session_state.get("feedback_completed_at_utc") or utc_now()
    optional_items = (session_state.get("consent") or {}).get("optional_items", {})
    comment_source = session_state.get("feedback_open_comment", "") if optional_items.get("open_text_feedback_storage") else ""
    record.update(
        {
            "record_updated_at_utc": utc_now(),
            "feedback_completed_at_utc": completed_at,
            "feedback_submitted": True,
            "feedback_ease_of_use": session_state.get("feedback_ease_of_use", ""),
            "feedback_question_clarity": session_state.get("feedback_question_clarity", ""),
            "feedback_report_clarity": session_state.get("feedback_report_clarity", ""),
            "feedback_comfort": session_state.get("feedback_comfort", ""),
            "feedback_trust": session_state.get("feedback_trust", ""),
            "feedback_usefulness": session_state.get("feedback_usefulness", ""),
            "feedback_recommend": session_state.get("feedback_recommend", ""),
            "feedback_open_comment_redacted": sanitize_free_text(comment_source),
            "record_status": "feedback_updated",
        }
    )
    return record


def _ensure_headers(worksheet: Any, headers: list[str]) -> None:
    existing = [worksheet.cell(row=1, column=index).value for index in range(1, max(worksheet.max_column, len(headers)) + 1)]
    if not existing or existing[0] is None:
        for column_index, header in enumerate(headers, start=1):
            worksheet.cell(row=1, column=column_index).value = header
        return
    missing = [header for header in headers if header not in existing]
    if missing:
        for header in missing:
            worksheet.cell(row=1, column=worksheet.max_column + 1).value = header


def upsert_by_session_id(worksheet: Any, record: dict[str, Any], headers: list[str]) -> None:
    _ensure_headers(worksheet, headers)
    current_headers = [worksheet.cell(row=1, column=index).value for index in range(1, worksheet.max_column + 1)]
    row_index = None
    for index in range(2, worksheet.max_row + 1):
        if str(worksheet.cell(row=index, column=1).value) == str(record.get("session_id", "")):
            row_index = index
            break
    values_by_header = dict(zip(headers, _normalise_row(record, headers)))
    if row_index is None:
        row_index = worksheet.max_row + 1
    for column_index, header in enumerate(current_headers, start=1):
        worksheet.cell(row=row_index, column=column_index).value = values_by_header.get(header, "")


def save_to_local_xlsx(record: dict[str, Any], worksheet_name: str = "screening_records") -> bool:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    headers = get_record_headers()
    if RECORD_XLSX.exists():
        workbook = load_workbook(RECORD_XLSX)
    else:
        workbook = Workbook()
        workbook.active.title = worksheet_name
    if worksheet_name not in workbook.sheetnames:
        worksheet = workbook.create_sheet(worksheet_name)
    else:
        worksheet = workbook[worksheet_name]
    upsert_by_session_id(worksheet, record, headers)
    if "save_audit" not in workbook.sheetnames:
        audit = workbook.create_sheet("save_audit")
        for column_index, header in enumerate(get_audit_headers(), start=1):
            audit.cell(row=1, column=column_index).value = header
    workbook.save(RECORD_XLSX)
    workbook.close()
    return True


def _get_streamlit_secret(name: str, default: Any = None) -> Any:
    try:
        import streamlit as st

        return st.secrets.get(name, default)
    except Exception:
        return default


def _google_client() -> Any | None:
    _load_env()
    if not _env_bool("ENABLE_GOOGLE_SHEETS_SAVE", False):
        return None
    try:
        import gspread
        from google.oauth2.service_account import Credentials
    except Exception:
        return None

    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive",
    ]
    credentials_payload = _get_streamlit_secret("google_service_account")
    if not credentials_payload:
        credentials_payload = os.getenv("GOOGLE_SERVICE_ACCOUNT_JSON", "")
    try:
        if isinstance(credentials_payload, str) and credentials_payload.strip():
            credentials_payload = json.loads(credentials_payload)
        if isinstance(credentials_payload, dict) and credentials_payload:
            credentials = Credentials.from_service_account_info(credentials_payload, scopes=scopes)
            return gspread.authorize(credentials)
    except Exception:
        return None

    service_account_file = os.getenv("GOOGLE_SERVICE_ACCOUNT_FILE", "").strip()
    if service_account_file:
        try:
            credentials = Credentials.from_service_account_file(service_account_file, scopes=scopes)
            return gspread.authorize(credentials)
        except Exception:
            return None
    return None


def save_to_google_sheet(record: dict[str, Any], worksheet_name: str = "screening_records") -> bool:
    client = _google_client()
    if client is None:
        raise RuntimeError("Google Sheets save is not configured.")
    sheet_id = os.getenv("GOOGLE_SHEET_ID", DEFAULT_SHEET_ID)
    spreadsheet = client.open_by_key(sheet_id)
    headers = get_record_headers()
    try:
        worksheet = spreadsheet.worksheet(worksheet_name)
    except Exception:
        worksheet = spreadsheet.add_worksheet(title=worksheet_name, rows=1000, cols=len(headers))

    existing_values = worksheet.get_all_values()
    if not existing_values:
        worksheet.append_row(headers, value_input_option="USER_ENTERED")
        existing_values = [headers]
    current_headers = existing_values[0]
    missing_headers = [header for header in headers if header not in current_headers]
    if missing_headers:
        current_headers = current_headers + missing_headers
        worksheet.update("1:1", [current_headers])

    target_row = None
    for index, row in enumerate(existing_values[1:], start=2):
        if row and row[0] == str(record.get("session_id", "")):
            target_row = index
            break
    row_values = [record.get(header, "") for header in current_headers]
    if target_row:
        worksheet.update(f"A{target_row}", [row_values], value_input_option="USER_ENTERED")
    else:
        worksheet.append_row(row_values, value_input_option="USER_ENTERED")
    return True


def save_to_google_apps_script(record: dict[str, Any], worksheet_name: str = "screening_records") -> bool:
    _load_env()
    if not _env_bool("ENABLE_GOOGLE_APPS_SCRIPT_SAVE", False):
        raise RuntimeError("Google Apps Script save is not configured.")
    web_app_url = os.getenv("GOOGLE_APPS_SCRIPT_WEB_APP_URL", "").strip()
    shared_secret = os.getenv("GOOGLE_APPS_SCRIPT_SHARED_SECRET", "").strip()
    if not web_app_url or not shared_secret:
        raise RuntimeError("Google Apps Script URL or shared secret is missing.")

    response = requests.post(
        web_app_url,
        json={
            "secret": shared_secret,
            "spreadsheet_id": os.getenv("GOOGLE_SHEET_ID", DEFAULT_SHEET_ID),
            "worksheet_name": worksheet_name,
            "headers": get_audit_headers() if worksheet_name == "save_audit" else get_record_headers(),
            "record": record,
        },
        timeout=20,
    )
    response.raise_for_status()
    payload = response.json()
    if not payload.get("success"):
        raise RuntimeError(str(payload.get("error", "Apps Script save failed.")))
    return True


def _google_save_enabled() -> bool:
    _load_env()
    return _env_bool("ENABLE_GOOGLE_APPS_SCRIPT_SAVE", False) or _env_bool("ENABLE_GOOGLE_SHEETS_SAVE", False)


def save_to_google_record(record: dict[str, Any], worksheet_name: str = "screening_records") -> bool:
    _load_env()
    errors: list[str] = []
    if _env_bool("ENABLE_GOOGLE_APPS_SCRIPT_SAVE", False):
        try:
            return save_to_google_apps_script(record, worksheet_name)
        except Exception as error:
            errors.append(_safe_error(error))
    if _env_bool("ENABLE_GOOGLE_SHEETS_SAVE", False):
        try:
            return save_to_google_sheet(record, worksheet_name)
        except Exception as error:
            errors.append(_safe_error(error))
    raise RuntimeError("; ".join(errors) or "Google save is not configured.")


def _safe_error(error: Exception | str) -> str:
    text = str(error)
    text = EMAIL_RE.sub("[redacted-email]", text)
    text = URL_RE.sub("[redacted-url]", text)
    return text[:160]


def _append_local_audit(session_id: str, target: str, success: bool, error_message: str = "") -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    if RECORD_XLSX.exists():
        workbook = load_workbook(RECORD_XLSX)
    else:
        workbook = Workbook()
        workbook.active.title = "screening_records"
        for column_index, header in enumerate(get_record_headers(), start=1):
            workbook["screening_records"].cell(row=1, column=column_index).value = header
    if "save_audit" not in workbook.sheetnames:
        worksheet = workbook.create_sheet("save_audit")
        for column_index, header in enumerate(get_audit_headers(), start=1):
            worksheet.cell(row=1, column=column_index).value = header
    else:
        worksheet = workbook["save_audit"]
        if worksheet.max_row == 1 and worksheet.cell(row=1, column=1).value is None:
            for column_index, header in enumerate(get_audit_headers(), start=1):
                worksheet.cell(row=1, column=column_index).value = header
    worksheet.append([utc_now(), session_id, target, bool(success), _safe_error(error_message)])
    workbook.save(RECORD_XLSX)
    workbook.close()


def log_save_audit(session_id: str, target: str, success: bool, error_message: str = "") -> None:
    try:
        _append_local_audit(session_id, target, success, error_message)
    except Exception:
        pass
    try:
        if _env_bool("ENABLE_GOOGLE_APPS_SCRIPT_SAVE", False):
            save_to_google_apps_script(
                {
                    "session_id": session_id,
                    "timestamp_utc": utc_now(),
                    "save_target": target,
                    "success": bool(success),
                    "error_message": _safe_error(error_message),
                },
                "save_audit",
            )
            return
        client = _google_client()
        if client is None:
            return
        spreadsheet = client.open_by_key(os.getenv("GOOGLE_SHEET_ID", DEFAULT_SHEET_ID))
        try:
            worksheet = spreadsheet.worksheet("save_audit")
        except Exception:
            worksheet = spreadsheet.add_worksheet(title="save_audit", rows=1000, cols=len(get_audit_headers()))
            worksheet.append_row(get_audit_headers(), value_input_option="USER_ENTERED")
        worksheet.append_row([utc_now(), session_id, target, bool(success), _safe_error(error_message)], value_input_option="USER_ENTERED")
    except Exception:
        pass


def _save_record(record: dict[str, Any], worksheet_name: str, status: str) -> dict[str, Any]:
    _load_env()
    result = {
        "local_success": False,
        "google_success": False,
        "google_target": "",
        "errors": [],
        "record_path": str(RECORD_XLSX),
    }
    google_enabled = _google_save_enabled()
    local_enabled = _env_bool("ENABLE_LOCAL_XLSX_BACKUP", True)
    record = dict(record)
    record["record_status"] = status

    if local_enabled:
        try:
            save_to_local_xlsx(record, worksheet_name)
            result["local_success"] = True
            log_save_audit(record.get("session_id", ""), "local_xlsx", True)
        except Exception as error:
            result["errors"].append("Local Excel backup could not be saved. Close the workbook if it is open and try again.")
            log_save_audit(record.get("session_id", ""), "local_xlsx", False, _safe_error(error))
    else:
        result["errors"].append("Local Excel backup is disabled.")

    if google_enabled:
        try:
            record["save_target_local_success"] = result["local_success"]
            save_to_google_record(record, worksheet_name)
            result["google_success"] = True
            result["google_target"] = "google_apps_script" if _env_bool("ENABLE_GOOGLE_APPS_SCRIPT_SAVE", False) else "google_sheet"
            log_save_audit(record.get("session_id", ""), result["google_target"], True)
        except Exception as error:
            result["errors"].append("Google Sheet save failed; local Excel backup has been saved.")
            log_save_audit(record.get("session_id", ""), "google_apps_script_or_sheet", False, _safe_error(error))
    else:
        result["errors"].append("Google Sheet save is not configured; local Excel backup has been saved.")

    record["save_target_local_success"] = result["local_success"]
    record["save_target_google_success"] = result["google_success"]
    if not result["google_success"] and result["local_success"]:
        record["record_status"] = "save_failed_local_only"
    if status == "feedback_updated" and result["local_success"]:
        record["record_status"] = "feedback_updated"

    if result["local_success"]:
        try:
            save_to_local_xlsx(record, worksheet_name)
            if worksheet_name == "feedback_records":
                save_to_local_xlsx(record, "screening_records")
        except Exception:
            pass
    if result["google_success"]:
        try:
            save_to_google_record(record, worksheet_name)
            if worksheet_name == "feedback_records":
                save_to_google_record(record, "screening_records")
        except Exception:
            pass
    return result


def save_screening_record(record: dict[str, Any]) -> dict[str, Any]:
    return _save_record(record, "screening_records", "screening_saved")


def save_feedback_record(record: dict[str, Any]) -> dict[str, Any]:
    return _save_record(record, "feedback_records", "feedback_updated")
