from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path
from uuid import uuid4

import streamlit as st


CONSENT_TEXT = """Consent to Participate in Preliminary Mental Health Screening

This website provides an educational and preliminary mental health screening process. It includes a chatbot-based conversation and structured screening forms such as BC-QoL and MMSE.

This tool does not provide a medical diagnosis and does not replace consultation with a qualified healthcare professional.

Your responses may be used only for generating your screening result and for educational/research evaluation of this project. Please do not enter personal identifying information such as your full name, address, phone number, NHS number, or private medical record details.

The system will use an anonymous session ID to process your responses. Your identity will not be requested.

You may stop using the tool at any time.

If you feel unsafe, distressed, or at risk of harm, please contact emergency services or a qualified mental health professional immediately."""


CONSENT_ITEMS = {
    "screening_not_diagnosis": "I understand this tool is for educational and preliminary screening only and is not a diagnosis.",
    "no_personal_details": "I understand I should not enter my name, email, phone number, address, or other identifying details.",
    "research_data_storage": "I consent to my anonymous screening scores and feedback being stored for dissertation research analysis.",
    "google_sheet_storage": "I understand the anonymous research data may be saved in a secure Google Sheet and local Excel workbook accessible only to the researcher/supervisor.",
    "voluntary_participation": "I understand I can stop using the prototype at any time.",
}

OPTIONAL_CONSENT_ITEMS = {
    "open_text_feedback_storage": "I agree that my optional open-text feedback may be stored after automatic redaction of identifiable details.",
}


def build_default_consent_state() -> dict:
    return {
        "is_consented": False,
        "timestamp": None,
        "personal_details_collected": False,
        "confirmed_items": {key: False for key in CONSENT_ITEMS},
        "optional_items": {key: False for key in OPTIONAL_CONSENT_ITEMS},
        "demographics": {
            "age_group": "",
            "gender": "",
            "preferred_language": "English",
            "accessibility_support_required": False,
            "previous_screening_experience": "",
            "device_type": "",
        },
    }


def build_anonymous_session_id() -> str:
    return f"anonymous_{uuid4()}"


def consent_is_complete(consent_state: dict) -> bool:
    confirmed_items = consent_state.get("confirmed_items", {})
    return all(bool(confirmed_items.get(key)) for key in CONSENT_ITEMS)


def store_consent_excel(session: dict, output_path: str | Path = "outputs/consent_data.xlsx") -> None:
    try:
        from openpyxl import Workbook, load_workbook
    except ImportError:
        st.warning("Consent could not be written to Excel because `openpyxl` is not installed.")
        return

    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    consent_state = session.get("consent", {})
    demographics = consent_state.get("demographics", {})
    confirmed = consent_state.get("confirmed_items", {})
    optional = consent_state.get("optional_items", {})
    headers = [
        "session_id",
        "consent_timestamp",
        "is_consented",
        "personal_details_collected",
        "screening_not_diagnosis",
        "no_personal_details",
        "research_data_storage",
        "google_sheet_storage",
        "voluntary_participation",
        "open_text_feedback_storage",
        "age_group",
        "gender",
        "preferred_language",
        "accessibility_support_required",
        "previous_screening_experience",
        "device_type",
    ]
    row = [
        session.get("session_id", ""),
        consent_state.get("timestamp", ""),
        bool(consent_state.get("is_consented")),
        False,
        bool(confirmed.get("screening_not_diagnosis")),
        bool(confirmed.get("no_personal_details")),
        bool(confirmed.get("research_data_storage")),
        bool(confirmed.get("google_sheet_storage")),
        bool(confirmed.get("voluntary_participation")),
        bool(optional.get("open_text_feedback_storage")),
        demographics.get("age_group", ""),
        demographics.get("gender", ""),
        demographics.get("preferred_language", ""),
        bool(demographics.get("accessibility_support_required")),
        demographics.get("previous_screening_experience", ""),
        demographics.get("device_type", ""),
    ]

    if path.exists():
        workbook = load_workbook(path)
        worksheet = workbook.active
        existing_ids = {
            str(worksheet.cell(row=index, column=1).value)
            for index in range(2, worksheet.max_row + 1)
        }
        if str(session.get("session_id", "")) in existing_ids:
            workbook.close()
            return
    else:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Consent"
        worksheet.append(headers)

    worksheet.append(row)
    try:
        workbook.save(path)
    except PermissionError:
        st.warning(f"Consent data could not be saved because `{path}` is locked or unavailable.")
        return
    finally:
        workbook.close()
    session["consent_excel_path"] = str(path)


def render_consent_form(session: dict) -> bool:
    consent_state = session.setdefault("consent", build_default_consent_state())
    session.setdefault("session_id", build_anonymous_session_id())

    st.markdown('<main class="assessment-container consent-page">', unsafe_allow_html=True)
    st.markdown('<section class="assessment-card">', unsafe_allow_html=True)
    st.title("Consent Form")
    with st.expander("Read consent information", expanded=False):
        st.markdown(
            """
            <div class="consent-copy">
                <h3>Consent to Participate in Preliminary Mental Health Screening</h3>
                <p>This website provides an educational and preliminary mental health screening process. It includes a chatbot-based conversation and structured screening forms such as BC-QoL and MMSE.</p>
                <p>This tool does not provide a medical diagnosis and does not replace consultation with a qualified healthcare professional.</p>
                <p>Your responses may be used only for generating your screening result and for educational/research evaluation of this project. Please do not enter personal identifying information such as your full name, address, phone number, NHS number, or private medical record details.</p>
                <p>The system will use an anonymous session ID to process your responses. Your identity will not be requested.</p>
                <p>You may stop using the tool at any time.</p>
                <p>If you feel unsafe, distressed, or at risk of harm, please contact emergency services or a qualified mental health professional immediately.</p>
            </div>
            """,
            unsafe_allow_html=True,
        )

    with st.expander("Privacy and Research Data Notice - UK GDPR Summary", expanded=False):
        st.markdown(
            """
            <div class="consent-copy">
                <p><strong>Purpose:</strong> dissertation evaluation of an educational mental health screening prototype.</p>
                <p><strong>Data collected:</strong> anonymous session ID, age group, gender option, preferred language, device type, accessibility need flag, previous screening experience, screening completion status, emotion/risk summary, BC-QoL score, MMSE-style score, non-diagnostic outcome label, feedback ratings, and redacted optional comments.</p>
                <p><strong>Data not collected:</strong> name, email, phone number, address, exact date of birth, IP address, full raw chat transcript, or clinical record number.</p>
                <p><strong>Storage:</strong> local <code>outputs/research_screening_records.xlsx</code> plus the configured Google Sheet when credentials are enabled.</p>
                <p><strong>Access:</strong> researcher and authorised academic supervisor only.</p>
                <p><strong>Retention:</strong> Retained until dissertation assessment is complete and then deleted or anonymised further according to university policy. Confirm exact retention period with supervisor/DPO.</p>
                <p><strong>Your rights:</strong> you may ask about access, correction, or withdrawal where possible. Because records are anonymous, withdrawal may only be possible if you provide your anonymous session ID. You may also complain to the Information Commissioner's Office (ICO).</p>
                <p><strong>Automated decision-making:</strong> the app produces preliminary educational screening indicators only. No legal, clinical, financial, or similarly significant decision is made automatically.</p>
                <p><strong>Safety:</strong> if risk flags appear, the system provides signposting only and does not replace emergency or professional support.</p>
                <p>This prototype is designed to support UK GDPR-aligned research data handling; it must be reviewed before real participant testing.</p>
            </div>
            """,
            unsafe_allow_html=True,
        )

    st.markdown("### Required confirmation")
    confirmed_items = consent_state.setdefault("confirmed_items", {})
    for key, label in CONSENT_ITEMS.items():
        confirmed_items[key] = st.checkbox(
            label,
            value=bool(confirmed_items.get(key, False)),
            key=f"consent_{key}",
        )
    st.markdown("### Optional research feedback consent")
    optional_items = consent_state.setdefault("optional_items", {})
    for key, label in OPTIONAL_CONSENT_ITEMS.items():
        optional_items[key] = st.checkbox(
            label,
            value=bool(optional_items.get(key, False)),
            key=f"optional_consent_{key}",
        )

    st.markdown('<div class="participant-title">Participant details</div>', unsafe_allow_html=True)
    st.markdown('<div class="participant-section">', unsafe_allow_html=True)
    demographics = consent_state.setdefault("demographics", {})
    st.text_input("Anonymous ID", value=session["session_id"], disabled=True)
    demo_col1, demo_col2, demo_col3 = st.columns(3)
    age_options = ["", "18-25", "26-35", "36-45", "46-60", "60+"]
    demographics["age_group"] = demo_col1.selectbox(
        "Age group",
        age_options,
        index=age_options.index(demographics.get("age_group", "")) if demographics.get("age_group", "") in age_options else 0,
    )
    gender_options = ["", "Woman", "Man", "Non-binary", "Prefer not to say", "Other"]
    demographics["gender"] = demo_col2.selectbox(
        "Gender",
        gender_options,
        index=gender_options.index(demographics.get("gender", "")) if demographics.get("gender", "") in gender_options else 0,
    )
    language_options = ["English", "Other"]
    demographics["preferred_language"] = demo_col3.selectbox(
        "Preferred language",
        language_options,
        index=language_options.index(demographics.get("preferred_language", "English"))
        if demographics.get("preferred_language", "English") in language_options
        else 0,
    )
    extra_col1, extra_col2, extra_col3 = st.columns(3)
    accessibility_options = ["No", "Yes"]
    accessibility_value = "Yes" if demographics.get("accessibility_support_required") else "No"
    demographics["accessibility_support_required"] = (
        extra_col1.selectbox(
            "Accessibility support required",
            accessibility_options,
            index=accessibility_options.index(accessibility_value),
        )
        == "Yes"
    )
    previous_options = ["", "No", "Yes", "Prefer not to say"]
    demographics["previous_screening_experience"] = extra_col2.selectbox(
        "Used a screening tool before?",
        previous_options,
        index=previous_options.index(demographics.get("previous_screening_experience", ""))
        if demographics.get("previous_screening_experience", "") in previous_options
        else 0,
    )
    device_options = ["", "Desktop or laptop", "Tablet", "Mobile phone"]
    demographics["device_type"] = extra_col3.selectbox(
        "Device used today",
        device_options,
        index=device_options.index(demographics.get("device_type", ""))
        if demographics.get("device_type", "") in device_options
        else 0,
    )
    st.markdown("</div>", unsafe_allow_html=True)

    ready = consent_is_complete(consent_state)
    if st.button("Continue to screening", type="primary", disabled=not ready, use_container_width=True):
        consent_state["is_consented"] = True
        consent_state["timestamp"] = datetime.now(timezone.utc).isoformat()
        consent_state["personal_details_collected"] = False
        session["consent"] = consent_state
        store_consent_excel(session)
        st.rerun()

    if not ready:
        st.warning("Select all required consent checkboxes to continue.")
    st.markdown("</section></main>", unsafe_allow_html=True)
    return bool(consent_state.get("is_consented") and consent_is_complete(consent_state))
